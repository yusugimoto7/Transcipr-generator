"""Build the bilingual price-list workbook from pricelist_data.py."""

import pathlib

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from pricelist_data import CATEGORIES, CUSTOM_RULES, GOV, SERVICES

OUT = pathlib.Path(__file__).resolve().parent / "Sugimoto Group - Price List 2026.xlsx"

PLUM = "714B67"
HEAD = PatternFill("solid", fgColor=PLUM)
CAT = PatternFill("solid", fgColor="EFE6EC")
WHITE = Font(color="FFFFFF", bold=True)
BOLD = Font(bold=True)
THIN = Side(style="thin", color="D9D4DA")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
WRAP = Alignment(wrap_text=True, vertical="top")
WRAP_RTL = Alignment(wrap_text=True, vertical="top", horizontal="right", readingOrder=2)


def money(v, cur="CAD"):
    return f"{v:,.2f} {cur}".replace(".00 ", " ") if v else "—"


def header(ws, cols, widths):
    ws.append(cols)
    for i, w in enumerate(widths, start=1):
        c = ws.cell(row=1, column=i)
        c.fill, c.font, c.alignment, c.border = HEAD, WHITE, WRAP, BORDER
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 34


def style_row(ws, row, rtl_cols):
    for i in range(1, ws.max_column + 1):
        c = ws.cell(row=row, column=i)
        c.border = BORDER
        c.alignment = WRAP_RTL if i in rtl_cols else WRAP


def services_sheet(ws, currency):
    header(ws, ["Code", "Service", "خدمت", f"Principal ({currency})", "Add-ons", "خدمات افزوده",
                "Government fees", "Payment terms", "شرایط پرداخت", "Notes", "توضیحات"],
           [12, 34, 30, 14, 30, 30, 30, 48, 48, 30, 30])
    rtl = {3, 6, 9, 11}
    for cat, (cat_en, cat_fa) in CATEGORIES.items():
        rows = [(k, s) for k, s in SERVICES.items()
                if s["cat"] == cat and s.get("currency", "CAD") == currency]
        if not rows:
            continue
        ws.append([f"{cat_en}  ·  {cat_fa}"])
        r = ws.max_row
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=11)
        ws.cell(row=r, column=1).fill = CAT
        ws.cell(row=r, column=1).font = BOLD
        for k, s in rows:
            addons_en = "\n".join(f"{lab.split('|')[0]}: {money(p, currency)}"
                                  for lab, p in s.get("addons", {}).items())
            addons_fa = "\n".join(f"{lab.split('|')[1]}: {money(p, currency)}"
                                  for lab, p in s.get("addons", {}).items())
            if s.get("components"):
                addons_en = "\n".join(f"{en}: {money(p)}" for _, en, _, p in s["components"])
                addons_fa = "\n".join(f"{fa}: {money(p)}" for _, _, fa, p in s["components"])
            gov = "\n".join(f"{GOV[g][0].replace('Government fee – ', '')}: {money(GOV[g][2])}"
                            for g in s.get("gov", []))
            ws.append([k, s["en"], s["fa"], s.get("price") or None, addons_en, addons_fa, gov,
                       s.get("terms_en", ""), s.get("terms_fa", ""), s.get("notes_en", ""), s.get("notes_fa", "")])
            style_row(ws, ws.max_row, rtl)
            ws.cell(row=ws.max_row, column=4).number_format = "#,##0"


def gov_sheet(ws):
    header(ws, ["Code", "Government fee", "هزینه دولتی", "Amount (CAD)"], [14, 52, 46, 14])
    for k, (en, fa, p) in GOV.items():
        ws.append([k, en, fa, p])
        style_row(ws, ws.max_row, {3})
        ws.cell(row=ws.max_row, column=4).number_format = "#,##0.00"


def rules_sheet(ws):
    header(ws, ["Rule", "قانون", "How it works in Odoo", "نحوه کار در اودو"], [26, 26, 60, 60])
    for en, fa, how_en, how_fa in CUSTOM_RULES:
        ws.append([en, fa, how_en, how_fa])
        style_row(ws, ws.max_row, {2, 4})


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Canada · کانادا"
    services_sheet(ws, "CAD")
    services_sheet(wb.create_sheet("Europe · اروپا"), "EUR")
    gov_sheet(wb.create_sheet("Government fees · هزینه دولتی"))
    rules_sheet(wb.create_sheet("Custom pricing · قیمت سفارشی"))
    wb.save(OUT)
    print(f"wrote {OUT}")


if __name__ == "__main__":
    build()
